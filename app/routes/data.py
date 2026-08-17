import io
import json
import csv
import openpyxl
from datetime import date, datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, send_file
from flask_login import login_required, current_user
from app import db
from app.models import Subscription, Payment, Category, PaymentMethod
from app.utils import (
    parse_date,
    generate_subscriptions_csv,
    generate_payments_csv,
    generate_excel_workbook
)

bp = Blueprint('data', __name__)


@bp.route('/export', methods=['GET', 'POST'])
@login_required
def export_data():
    """Export data to JSON, CSV, or Excel (.xlsx)."""
    if request.method == 'POST':
        export_format = request.form.get('format', 'json').lower()
        export_categories = request.form.get('categories') == 'on'
        export_payment_methods = request.form.get('payment_methods') == 'on'
        export_subscriptions = request.form.get('subscriptions') == 'on'
        export_payments = request.form.get('payments') == 'on'

        categories = Category.query.filter_by(user_id=current_user.id).all() if export_categories else []
        payment_methods = PaymentMethod.query.filter_by(user_id=current_user.id).all() if export_payment_methods else []
        subscriptions = Subscription.query.filter_by(user_id=current_user.id).all() if export_subscriptions else []
        payments = Payment.query.filter_by(user_id=current_user.id).all() if export_payments else []

        timestamp = date.today().isoformat()

        # Excel Export
        if export_format == 'excel':
            excel_data = generate_excel_workbook(categories, payment_methods, subscriptions, payments)
            return send_file(
                io.BytesIO(excel_data),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f"subscription_tracker_export_{timestamp}.xlsx"
            )

        # CSV Export (Subscriptions default)
        elif export_format == 'csv':
            csv_data = generate_subscriptions_csv(subscriptions)
            return Response(
                csv_data,
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename=subscriptions_{timestamp}.csv'}
            )

        # JSON Export
        else:
            data = {
                'export_date': datetime.utcnow().isoformat(),
                'version': '2.0',
                'user': current_user.username
            }

            if export_categories:
                data['categories'] = [c.to_dict() for c in categories]

            if export_payment_methods:
                data['payment_methods'] = [m.to_dict() for m in payment_methods]

            if export_subscriptions:
                data['subscriptions'] = [
                    {
                        'name': s.name,
                        'category_name': s.category.name if s.category else None,
                        'payment_method_name': s.payment_method.name if s.payment_method else None,
                        'amount': s.amount,
                        'currency': s.currency,
                        'billing_cycle': s.billing_cycle,
                        'next_due_date': s.next_due_date.isoformat() if s.next_due_date else None,
                        'url': s.url,
                        'notes': s.notes,
                        'tags': s.tags,
                        'icon': s.icon,
                        'is_active': s.is_active,
                        'is_variable': s.is_variable
                    }
                    for s in subscriptions
                ]

            if export_payments:
                data['payments'] = [
                    {
                        'subscription_name': p.subscription.name if p.subscription else None,
                        'payment_method_name': p.payment_method.name if p.payment_method else None,
                        'amount': p.amount,
                        'original_amount': p.original_amount,
                        'currency': p.currency,
                        'paid_date': p.paid_date.isoformat() if p.paid_date else None,
                        'notes': p.notes
                    }
                    for p in payments
                ]

            json_str = json.dumps(data, indent=2)
            filename = f"subscription_tracker_export_{timestamp}.json"
            return Response(
                json_str,
                mimetype='application/json',
                headers={'Content-Disposition': f'attachment; filename={filename}'}
            )

    stats = {
        'categories': Category.query.filter_by(user_id=current_user.id).count(),
        'payment_methods': PaymentMethod.query.filter_by(user_id=current_user.id).count(),
        'subscriptions': Subscription.query.filter_by(user_id=current_user.id).count(),
        'payments': Payment.query.filter_by(user_id=current_user.id).count()
    }
    return render_template('export.html', stats=stats)


@bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_data():
    """Import data from JSON, CSV, or Excel (.xlsx) file."""
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('No file selected', 'error')
            return redirect(url_for('data.import_data'))

        filename = file.filename.lower()
        skip_existing = request.form.get('skip_existing') == 'on'
        imported = {'categories': 0, 'payment_methods': 0, 'subscriptions': 0, 'payments': 0}

        try:
            # JSON Import
            if filename.endswith('.json'):
                data = json.load(file)

                # Categories
                if 'categories' in data:
                    for cat_data in data['categories']:
                        cat_name = cat_data.get('name')
                        if not cat_name:
                            continue
                        existing = Category.query.filter_by(user_id=current_user.id, name=cat_name).first()
                        if existing:
                            if not skip_existing:
                                existing.color = cat_data.get('color', existing.color)
                                existing.icon = cat_data.get('icon', existing.icon)
                        else:
                            cat = Category(
                                user_id=current_user.id,
                                name=cat_name,
                                color=cat_data.get('color', '#6b7280'),
                                icon=cat_data.get('icon', 'box')
                            )
                            db.session.add(cat)
                        imported['categories'] += 1
                    db.session.commit()

                # Payment Methods
                if 'payment_methods' in data:
                    for pm_data in data['payment_methods']:
                        pm_name = pm_data.get('name')
                        if not pm_name:
                            continue
                        existing = PaymentMethod.query.filter_by(user_id=current_user.id, name=pm_name).first()
                        if existing:
                            if not skip_existing:
                                existing.method_type = pm_data.get('method_type', existing.method_type)
                                existing.identifier = pm_data.get('identifier', existing.identifier)
                                existing.color = pm_data.get('color', existing.color)
                                existing.icon = pm_data.get('icon', existing.icon)
                        else:
                            pm = PaymentMethod(
                                user_id=current_user.id,
                                name=pm_name,
                                method_type=pm_data.get('method_type', 'bank'),
                                identifier=pm_data.get('identifier'),
                                color=pm_data.get('color', '#6b7280'),
                                icon=pm_data.get('icon', 'credit-card'),
                                is_default=pm_data.get('is_default', False)
                            )
                            db.session.add(pm)
                        imported['payment_methods'] += 1
                    db.session.commit()

                # Subscriptions
                if 'subscriptions' in data:
                    for sub_data in data['subscriptions']:
                        sub_name = sub_data.get('name')
                        if not sub_name:
                            continue
                        existing = Subscription.query.filter_by(user_id=current_user.id, name=sub_name).first()
                        if existing and skip_existing:
                            continue

                        category = None
                        if sub_data.get('category_name'):
                            category = Category.query.filter_by(user_id=current_user.id, name=sub_data['category_name']).first()

                        payment_method = None
                        if sub_data.get('payment_method_name'):
                            payment_method = PaymentMethod.query.filter_by(user_id=current_user.id, name=sub_data['payment_method_name']).first()

                        due_date = parse_date(sub_data.get('next_due_date'))

                        if existing:
                            existing.category_id = category.id if category else existing.category_id
                            existing.payment_method_id = payment_method.id if payment_method else existing.payment_method_id
                            existing.amount = float(sub_data.get('amount', existing.amount))
                            existing.currency = sub_data.get('currency', existing.currency)
                            existing.billing_cycle = sub_data.get('billing_cycle', existing.billing_cycle)
                            existing.next_due_date = due_date
                            existing.url = sub_data.get('url', existing.url)
                            existing.notes = sub_data.get('notes', existing.notes)
                            existing.tags = sub_data.get('tags', existing.tags)
                            existing.icon = sub_data.get('icon', existing.icon)
                            existing.is_active = sub_data.get('is_active', existing.is_active)
                            existing.is_variable = sub_data.get('is_variable', existing.is_variable)
                        else:
                            sub = Subscription(
                                user_id=current_user.id,
                                name=sub_name,
                                category_id=category.id if category else None,
                                payment_method_id=payment_method.id if payment_method else None,
                                amount=float(sub_data.get('amount', 0)),
                                currency=sub_data.get('currency', 'TRY'),
                                billing_cycle=sub_data.get('billing_cycle', 'monthly'),
                                next_due_date=due_date,
                                url=sub_data.get('url'),
                                notes=sub_data.get('notes'),
                                tags=sub_data.get('tags'),
                                icon=sub_data.get('icon', 'receipt'),
                                is_active=sub_data.get('is_active', True),
                                is_variable=sub_data.get('is_variable', False)
                            )
                            db.session.add(sub)
                        imported['subscriptions'] += 1
                    db.session.commit()

            # CSV Import (Subscriptions)
            elif filename.endswith('.csv'):
                content = file.read().decode('utf-8-sig', errors='replace')
                reader = csv.DictReader(io.StringIO(content))
                for row in reader:
                    name = row.get('Name') or row.get('name')
                    if not name:
                        continue

                    existing = Subscription.query.filter_by(user_id=current_user.id, name=name).first()
                    if existing and skip_existing:
                        continue

                    cat_name = row.get('Category') or row.get('category')
                    category = None
                    if cat_name:
                        category = Category.query.filter_by(user_id=current_user.id, name=cat_name).first()
                        if not category:
                            category = Category(user_id=current_user.id, name=cat_name, color='#6366f1', icon='tag')
                            db.session.add(category)
                            db.session.flush()

                    pm_name = row.get('Payment Method') or row.get('payment_method')
                    pm = None
                    if pm_name:
                        pm = PaymentMethod.query.filter_by(user_id=current_user.id, name=pm_name).first()

                    amount_str = row.get('Amount') or row.get('amount') or '0'
                    try:
                        amount = float(amount_str)
                    except ValueError:
                        amount = 0.0

                    due_date = parse_date(row.get('Next Due Date') or row.get('next_due_date'))

                    if existing:
                        existing.amount = amount
                        existing.currency = row.get('Currency') or existing.currency
                        existing.billing_cycle = row.get('Billing Cycle') or existing.billing_cycle
                        existing.next_due_date = due_date
                    else:
                        sub = Subscription(
                            user_id=current_user.id,
                            name=name,
                            category_id=category.id if category else None,
                            payment_method_id=pm.id if pm else None,
                            amount=amount,
                            currency=row.get('Currency') or 'TRY',
                            billing_cycle=row.get('Billing Cycle') or 'monthly',
                            next_due_date=due_date,
                            url=row.get('URL') or None,
                            notes=row.get('Notes') or None,
                            tags=row.get('Tags') or None,
                            is_active=True
                        )
                        db.session.add(sub)
                    imported['subscriptions'] += 1
                db.session.commit()

            # Excel Import (.xlsx)
            elif filename.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file, data_only=True)
                if "Subscriptions" in wb.sheetnames:
                    ws = wb["Subscriptions"]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) > 1:
                        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
                        for row in rows[1:]:
                            if not row or not any(row):
                                continue
                            row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                            name = row_dict.get('name')
                            if not name:
                                continue

                            existing = Subscription.query.filter_by(user_id=current_user.id, name=str(name)).first()
                            if existing and skip_existing:
                                continue

                            cat_name = row_dict.get('category')
                            cat = Category.query.filter_by(user_id=current_user.id, name=str(cat_name)).first() if cat_name else None

                            pm_name = row_dict.get('payment method')
                            pm = PaymentMethod.query.filter_by(user_id=current_user.id, name=str(pm_name)).first() if pm_name else None

                            amount = float(row_dict.get('amount') or 0.0)
                            due_date = parse_date(row_dict.get('next due date'))

                            if existing:
                                existing.amount = amount
                                existing.next_due_date = due_date
                            else:
                                sub = Subscription(
                                    user_id=current_user.id,
                                    name=str(name),
                                    category_id=cat.id if cat else None,
                                    payment_method_id=pm.id if pm else None,
                                    amount=amount,
                                    currency=str(row_dict.get('currency') or 'TRY'),
                                    billing_cycle=str(row_dict.get('billing cycle') or 'monthly'),
                                    next_due_date=due_date,
                                    url=str(row_dict.get('url')) if row_dict.get('url') else None,
                                    notes=str(row_dict.get('notes')) if row_dict.get('notes') else None,
                                    tags=str(row_dict.get('tags')) if row_dict.get('tags') else None,
                                    is_active=True
                                )
                                db.session.add(sub)
                            imported['subscriptions'] += 1
                        db.session.commit()

            flash(f"Import successful! Subscriptions: {imported['subscriptions']}, Categories: {imported['categories']}, Payment Methods: {imported['payment_methods']}", 'success')
            return redirect(url_for('subscriptions.index'))

        except Exception as e:
            db.session.rollback()
            flash(f"Error during import: {e}", 'error')
            return redirect(url_for('data.import_data'))

    return render_template('import.html')
