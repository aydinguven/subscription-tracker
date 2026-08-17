"""Shared utility functions for date parsing, formatting, and file export/import."""

import csv
import io
from datetime import date, datetime
import openpyxl


def parse_date(date_input):
    """Parse date string from various formats into a datetime.date object.

    Supports:
    - Already a date or datetime instance
    - ISO format: YYYY-MM-DD
    - DD Mon YY / DD Mon YYYY: '06 Feb 26', '06 Feb 2026'
    - DD/MM/YYYY, DD-MM-YYYY, MM/DD/YYYY
    """
    if not date_input:
        return None

    if isinstance(date_input, datetime):
        return date_input.date()
    if isinstance(date_input, date):
        return date_input

    date_string = str(date_input).strip()
    if not date_string:
        return None

    # Try ISO format first
    try:
        return date.fromisoformat(date_string)
    except ValueError:
        pass

    formats = [
        '%Y-%m-%d',
        '%d %b %y',
        '%d %b %Y',
        '%d/%m/%Y',
        '%d/%m/%y',
        '%d-%m-%Y',
        '%d-%m-%y',
        '%m/%d/%Y',
        '%m/%d/%y',
        '%Y/%m/%d',
    ]

    for fmt in formats:
        try:
            return datetime.strptime(date_string, fmt).date()
        except ValueError:
            continue

    raise ValueError(
        f"Could not parse date: '{date_string}'. "
        f"Expected formats: YYYY-MM-DD, DD/MM/YYYY, DD Mon YYYY"
    )


def generate_subscriptions_csv(subscriptions):
    """Generate CSV string from subscriptions list."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Name', 'Category', 'Payment Method', 'Amount', 'Currency',
        'Billing Cycle', 'Next Due Date', 'URL', 'Tags', 'Is Variable', 'Is Active', 'Notes'
    ])

    for s in subscriptions:
        writer.writerow([
            s.name,
            s.category.name if s.category else '',
            s.payment_method.name if s.payment_method else '',
            s.amount,
            s.currency,
            s.billing_cycle,
            s.next_due_date.isoformat() if s.next_due_date else '',
            s.url or '',
            s.tags or '',
            'Yes' if s.is_variable else 'No',
            'Yes' if s.is_active else 'No',
            s.notes or ''
        ])

    return output.getvalue()


def generate_payments_csv(payments):
    """Generate CSV string from payments list."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Subscription', 'Payment Method', 'Amount Paid', 'Original Amount',
        'Currency', 'Paid Date', 'Notes'
    ])

    for p in payments:
        writer.writerow([
            p.subscription.name if p.subscription else '',
            p.payment_method.name if p.payment_method else '',
            p.amount,
            p.original_amount if p.original_amount else '',
            p.currency,
            p.paid_date.isoformat() if p.paid_date else '',
            p.notes or ''
        ])

    return output.getvalue()


def generate_excel_workbook(categories, payment_methods, subscriptions, payments):
    """Generate Excel binary workbook (.xlsx) with separate sheets."""
    wb = openpyxl.Workbook()

    # Subscriptions Sheet
    ws_subs = wb.active
    ws_subs.title = "Subscriptions"
    ws_subs.append([
        'Name', 'Category', 'Payment Method', 'Amount', 'Currency',
        'Billing Cycle', 'Next Due Date', 'URL', 'Tags', 'Is Variable', 'Is Active', 'Notes'
    ])
    for s in subscriptions:
        ws_subs.append([
            s.name,
            s.category.name if s.category else '',
            s.payment_method.name if s.payment_method else '',
            s.amount,
            s.currency,
            s.billing_cycle,
            s.next_due_date.isoformat() if s.next_due_date else '',
            s.url or '',
            s.tags or '',
            'Yes' if s.is_variable else 'No',
            'Yes' if s.is_active else 'No',
            s.notes or ''
        ])

    # Payments Sheet
    ws_pay = wb.create_sheet(title="Payments")
    ws_pay.append([
        'Subscription', 'Payment Method', 'Amount Paid', 'Original Amount',
        'Currency', 'Paid Date', 'Notes'
    ])
    for p in payments:
        ws_pay.append([
            p.subscription.name if p.subscription else '',
            p.payment_method.name if p.payment_method else '',
            p.amount,
            p.original_amount if p.original_amount else '',
            p.currency,
            p.paid_date.isoformat() if p.paid_date else '',
            p.notes or ''
        ])

    # Categories Sheet
    ws_cat = wb.create_sheet(title="Categories")
    ws_cat.append(['Name', 'Color', 'Icon'])
    for c in categories:
        ws_cat.append([c.name, c.color, c.icon])

    # Payment Methods Sheet
    ws_pm = wb.create_sheet(title="Payment Methods")
    ws_pm.append(['Name', 'Method Type', 'Identifier', 'Color', 'Icon', 'Is Default'])
    for m in payment_methods:
        ws_pm.append([m.name, m.method_type, m.identifier or '', m.color, m.icon, 'Yes' if m.is_default else 'No'])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
